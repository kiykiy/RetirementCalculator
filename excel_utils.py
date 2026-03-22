"""
Excel/XLSX Utility Script
Covers: read, write, edit, convert
Libraries: openpyxl (low-level), pandas (data analysis)
"""

import openpyxl
import pandas as pd


# ─── READ ────────────────────────────────────────────────────────────────────

def read_excel(file_path, sheet_name=0):
    """Read an Excel file into a pandas DataFrame."""
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(df)
    return df


def read_all_sheets(file_path):
    """Read all sheets from an Excel file."""
    sheets = pd.read_excel(file_path, sheet_name=None)
    for name, df in sheets.items():
        print(f"\n--- Sheet: {name} ---")
        print(df)
    return sheets


# ─── WRITE ───────────────────────────────────────────────────────────────────

def write_excel(data: dict, file_path: str):
    """
    Write data to an Excel file.
    data: {"Sheet1": [{"col1": val, "col2": val}, ...], ...}
    """
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, rows in data.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Saved: {file_path}")


# ─── EDIT ────────────────────────────────────────────────────────────────────

def edit_cell(file_path, sheet_name, cell, value):
    """Edit a single cell in an existing Excel file."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    ws[cell] = value
    wb.save(file_path)
    print(f"Set {sheet_name}!{cell} = {value!r}")


def append_rows(file_path, sheet_name, rows: list):
    """Append rows to an existing sheet."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for row in rows:
        ws.append(row)
    wb.save(file_path)
    print(f"Appended {len(rows)} row(s) to {sheet_name}")


# ─── CONVERT ─────────────────────────────────────────────────────────────────

def xlsx_to_csv(xlsx_path, csv_path, sheet_name=0):
    """Convert an Excel sheet to CSV."""
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    df.to_csv(csv_path, index=False)
    print(f"Saved CSV: {csv_path}")


def csv_to_xlsx(csv_path, xlsx_path, sheet_name="Sheet1"):
    """Convert a CSV file to Excel."""
    df = pd.read_csv(csv_path)
    df.to_excel(xlsx_path, sheet_name=sheet_name, index=False)
    print(f"Saved XLSX: {xlsx_path}")


def xlsx_to_json(xlsx_path, sheet_name=0):
    """Convert an Excel sheet to JSON string."""
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    json_str = df.to_json(orient="records", indent=2)
    print(json_str)
    return json_str


# ─── EXAMPLE USAGE ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Write example
    write_excel(
        data={
            "Sales": [
                {"Name": "Alice", "Amount": 1000, "Month": "Jan"},
                {"Name": "Bob",   "Amount": 1500, "Month": "Jan"},
            ],
            "Summary": [
                {"Total": 2500}
            ]
        },
        file_path="example.xlsx"
    )

    # Read it back
    read_excel("example.xlsx", sheet_name="Sales")

    # Edit a cell
    edit_cell("example.xlsx", "Sales", "B2", 9999)

    # Append rows
    append_rows("example.xlsx", "Sales", [["Charlie", 2000, "Feb"]])

    # Convert to CSV
    xlsx_to_csv("example.xlsx", "sales.csv", sheet_name="Sales")

    # Convert to JSON
    xlsx_to_json("example.xlsx", sheet_name="Sales")
